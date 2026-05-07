import sys
import pandas as pd
import shutil
import xlwings as xw

from pathlib import Path
from openpyxl.utils import get_column_letter
from utils.date_utils import today_str
from email_pipeline import email_setup as config
from utils import pivot_utils
from email_pipeline.downloader import EmailDownloader
from email_pipeline.sender import load_sender, send_report, send_skip

DATA_DIR   = config.DATA_DIR
OUTPUT_DIR = config.OUTPUT_DIR
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def find_file(pattern: str):
    """Return the first file in DATA_DIR matching pattern. Raises FileNotFoundError if none found."""
    matches = list(DATA_DIR.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"No file matching '{pattern}' in {DATA_DIR}")
    return matches[0]

def provider_c_report():
    df = pd.read_excel(find_file("Provider C Report*.xlsx"))

    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    # Merge B~F into a single "Vendor Full Address" column at G position
    g_idx = list(df.columns).index(col_dict["G"])
    g_val = (df[col_dict["B"]].astype(str) + ", " +
             df[col_dict["C"]].astype(str) + ", " +
             df[col_dict["D"]].astype(str) + ", " +
             df[col_dict["E"]].astype(str) + ", " +
             df[col_dict["F"]].astype(str))
    df.insert(g_idx, "Vendor Full Address", g_val)

    df.drop(columns=[col_dict["B"], col_dict["C"], col_dict["D"],
                     col_dict["E"], col_dict["F"]], inplace=True)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    # Move I~N columns to C position
    i_to_n = df.loc[:, col_dict["I"]:col_dict["N"]].copy()
    df.drop(columns=i_to_n.columns.tolist(), inplace=True)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    c_idx = list(df.columns).index(col_dict["C"])
    for i, col in enumerate(i_to_n.columns):
        df.insert(c_idx + i, col, i_to_n[col].values)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    # Move M~N (PU Unix) to I position
    m_n = df.loc[:, col_dict["M"]:col_dict["N"]].copy()
    df.drop(columns=m_n.columns.tolist(), inplace=True)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    i_idx = list(df.columns).index(col_dict["I"])
    for i, col in enumerate(m_n.columns):
        df.insert(i_idx + i, col, m_n[col].values)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    # Insert 2 blank columns at L
    l_idx = list(df.columns).index(col_dict["L"])
    df.insert(l_idx, "New_L1", "")
    df.insert(l_idx + 1, "New_L2", "")
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    # Move Q (PU Unix date) → L, clear Q
    df[col_dict["L"]] = df[col_dict["Q"]].dt.strftime("%#m/%#d/%Y")
    df.rename(columns={col_dict["L"]: col_dict["Q"], col_dict["Q"]: "Temp_Q"}, inplace=True)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}
    df[col_dict["Q"]] = None

    # Insert 4 blank columns at O
    o_idx = list(df.columns).index(col_dict["O"])
    for i in range(4):
        df.insert(o_idx + i, f"New_O{i + 1}", "")
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    # Move X (ETA FD Unix) → S, clear X
    df[col_dict["S"]] = df[col_dict["X"]]
    df.rename(columns={col_dict["S"]: col_dict["X"], col_dict["X"]: "Temp_X"}, inplace=True)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}
    df[col_dict["X"]] = None

    df = df.loc[:, :col_dict["S"]]

    print("Provider C report processed")
    return df.loc[:, col_dict["A"]:col_dict["S"]]


def provider_d_report():
    df = pd.read_csv(find_file("Provider D Daily Report*.csv"), index_col=False)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    df = df.loc[:, :col_dict["T"]]
    df[col_dict["A"]] = "Provider D"

    df.drop(columns=[col_dict["C"]], inplace=True)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    # Split D column (full address) into separate address fields
    d_col = col_dict["D"]
    split_df = df[d_col].str.split(",", expand=True).apply(lambda x: x.str.strip())
    split_df.columns = ["ConsigneeAddress", "City", "State", "Country", "Zip"]

    d_idx = list(df.columns).index(d_col)
    df.drop(columns=[d_col], inplace=True)
    for i, col in enumerate(split_df.columns):
        df.insert(d_idx + i, col, split_df[col])
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    df.drop(columns=[col_dict["I"], col_dict["J"]], inplace=True)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    df.drop(columns=[col_dict["K"], col_dict["L"], col_dict["M"]], inplace=True)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    df.drop(columns=[col_dict["L"], col_dict["M"], col_dict["N"], col_dict["O"], col_dict["P"]], inplace=True)
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    m_idx = list(df.columns).index(col_dict["M"])
    df.insert(m_idx, "New_M", "")
    col_dict = {get_column_letter(i + 1): col for i, col in enumerate(df.columns)}

    df = df.loc[:, :col_dict["N"]]
    df[col_dict["K"]] = df[col_dict["K"]].fillna("SHIPMENT EN-ROUTE-TO-DEST")

    print("Provider D report processed")
    return df.loc[:, col_dict["A"]:col_dict["N"]]


def combine_excel():
    # Use template header row as the column reference
    template_cols = pd.read_excel(config.TEMPLATE_PATH, sheet_name="Daily Manifest", nrows=0).columns

    dfs = []
    for label, loader in [
        ("Provider A", lambda: pd.read_excel(find_file("Provider A Daily Exception List*.xlsx"))),
        ("Provider B", lambda: pd.read_excel(find_file("Provider B Daily Exception List*.xlsx")).iloc[:, 0:19]),
        ("Provider C", provider_c_report),
        ("Provider D", provider_d_report),
        ("Provider E", lambda: pd.read_excel(find_file("Provider E Daily Manifest*.xlsx")).iloc[:, 0:19]),
    ]:
        try:
            df = loader()
            df.columns = template_cols[:len(df.columns)]
            dfs.append(df)
        except FileNotFoundError:
            print(f"[SKIP] {label} not found")

    if not dfs:
        raise RuntimeError("No files available to consolidate")

    df_result = pd.concat(dfs, ignore_index=True)

    shutil.copy(config.TEMPLATE_PATH, OUTPUT_DIR / "Consolidated Daily Manifest.xlsx")

    app = xw.App(visible=False)
    wb = app.books.open(str(OUTPUT_DIR / "Consolidated Daily Manifest.xlsx"))

    ws = wb.sheets["Daily Manifest"]
    ws.range("A2").value = df_result.values.tolist()

    # Update pivot cache, refresh, and capture data
    pivot_data = pivot_utils.read_pivot(wb)

    wb.sheets["Pivot"].activate()
    xlsb_path = str(OUTPUT_DIR / "Consolidated Daily Manifest - {}.xlsb".format(today_str('e')))
    app.api.DisplayAlerts = False  # Suppress overwrite prompt
    wb.api.SaveAs(xlsb_path, FileFormat=50)
    app.api.DisplayAlerts = True
    wb.close()
    app.quit()
    print("Combine Completed!")

    print(f"Output file: {xlsb_path}")
    return Path(xlsb_path), pivot_data

if __name__ == "__main__":
    import os
    from dotenv import load_dotenv
    load_dotenv(config.ENV_PATH)

    pop3_user = os.environ.get(config.POP3_USER_ENV)
    pop3_pass = os.environ.get(config.POP3_PASS_ENV)
    result = EmailDownloader(pop3_user, pop3_pass).download()

    sender = load_sender()
    xlsb = OUTPUT_DIR / config.ATTACHMENT_FILENAME.format(today_str('e'))
    mode = sys.argv[1] if len(sys.argv) > 1 else "initial"

    has_after_1700 = any(
        int(Path(f).stem.rsplit('_', 1)[-1]) >= 1700
        for f in result.downloaded_files
    )
    has_after_1800 = any(
        int(Path(f).stem.rsplit('_', 1)[-1]) >= 1800
        for f in result.downloaded_files
    )

    if not xlsb.exists():
        output_file, pivot_data = combine_excel()
        send_report(sender, output_file, result.received_times, is_updated=False, pivot_data=pivot_data)
    elif mode == "second" and has_after_1700:
        output_file, pivot_data = combine_excel()
        send_report(sender, output_file, result.received_times, is_updated=True, pivot_data=pivot_data)
    elif mode == "third" and has_after_1800:
        output_file, pivot_data = combine_excel()
        send_report(sender, output_file, result.received_times, is_updated=True, pivot_data=pivot_data)
    else:
        send_skip(sender, result.downloaded_files)
